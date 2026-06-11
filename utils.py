import csv
import io
import sqlite3
from config import Config

_schema_cache: dict = {}

def load_schema_to_cache(db_path: str = None) -> dict:
    global _schema_cache
    db_path = db_path or Config.DATABASE_PATH

    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [r[0] for r in c.fetchall()]

    cache = {}
    for table in tables:
        c.execute(f"PRAGMA table_info({table})")
        columns = c.fetchall()

        try:
            c.execute(f"SELECT * FROM {table} LIMIT 3")
            samples = c.fetchall()
        except Exception:
            samples = []

        cache[table] = {
            "columns": [
                {"cid": col[0], "name": col[1], "type": col[2],
                 "notnull": col[3], "pk": col[5]}
                for col in columns
            ],
            "sample": [list(r) for r in samples],
        }

    conn.close()
    _schema_cache = cache
    print(f"[SCHEMA] Cache dimuat: {len(tables)} tabel")
    # print(cache)
    return cache

def get_schema_cache() -> dict:
    if not _schema_cache:
        load_schema_to_cache()
    return _schema_cache


def get_schema_for_display() -> dict:
    return {
        table: {"columns": info["columns"], "sample_count": len(info.get("sample", []))}
        for table, info in get_schema_cache().items()
    }

def execute_sql(sql: str, db_path: str = None) -> dict:
    db_path = db_path or Config.DATABASE_PATH
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute(sql)

        if c.description:
            columns = [d[0] for d in c.description]
            rows    = [list(r) for r in c.fetchall()]
            conn.close()
            return {"success": True, "columns": columns,
                    "rows": rows, "rowcount": len(rows), "error": None}
        else:
            rowcount = c.rowcount
            conn.commit()
            conn.close()
            return {"success": True, "columns": [], "rows": [],
                    "rowcount": rowcount, "error": None}

    except Exception as e:
        return {"success": False, "columns": [], "rows": [],
                "rowcount": 0, "error": str(e)}


def run_explain_query(sql: str, db_path: str = None) -> dict:
    return execute_sql(f"EXPLAIN QUERY PLAN {sql}", db_path)

def log_to_audit(username: str, role: str, pertanyaan: str,
                 sql_query: str, status: str, pesan: str = ""):
    try:
        conn = sqlite3.connect(Config.AUDIT_DB_PATH)
        conn.execute(
            "INSERT INTO audit_log (username, role, pertanyaan, sql_query, status, pesan) "
            "VALUES (?,?,?,?,?,?)",
            (username, role, pertanyaan, sql_query, status, pesan)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[AUDIT] Error: {e}")


def get_audit_logs(limit: int = 200) -> list:
    try:
        conn = sqlite3.connect(Config.AUDIT_DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM audit_log ORDER BY waktu DESC LIMIT ?", (limit,)
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception:
        return []


def export_to_csv(columns: list, rows: list) -> str:
    out = io.StringIO()
    w   = csv.writer(out)
    w.writerow(columns)
    w.writerows(rows)
    return out.getvalue()

def export_to_excel(columns: list, rows: list) -> bytes | None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hasil Query"

        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

        for ci, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    except ImportError:
        return None

def process_uploaded_db(filepath: str) -> dict:
    try:
        conn   = sqlite3.connect(filepath)
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]
        conn.close()
        if not tables:
            return {"success": False, "error": "Database kosong."}
        return {"success": True, "tables": tables}
    except Exception as e:
        return {"success": False, "error": str(e)}


def process_uploaded_csv(filepath: str, table_name: str, db_path: str) -> dict:
    try:
        import pandas as pd
        from sqlalchemy import create_engine

        df     = pd.read_csv(filepath)
        engine = create_engine(f"sqlite:///{db_path}")
        df.to_sql(table_name, engine, if_exists="replace", index=False)
        return {"success": True, "rows": len(df), "columns": list(df.columns)}
    except ImportError:
        return {"success": False, "error": "pandas tidak terinstall."}
    except Exception as e:
        return {"success": False, "error": str(e)}

def save_chat_message(username: str, role: str, content: str, sql_query: str = None, 
                      table_data: dict = None, visualization: dict = None, metadata: dict = None):
    import json
    try:
        conn = sqlite3.connect(Config.AUDIT_DB_PATH)
        
        # Convert dicts to JSON strings
        table_json = json.dumps(table_data) if table_data else None
        viz_json = json.dumps(visualization) if visualization else None
        meta_json = json.dumps(metadata) if metadata else None
        
        conn.execute(
            """INSERT INTO chat_history 
               (username, role, content, sql_query, table_data, visualization, metadata) 
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (username, role, content, sql_query, table_json, viz_json, meta_json)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[CHAT_HISTORY] Error saving message: {e}")


def get_chat_history(username: str, limit: int = 20) -> list:
    import json
    try:
        conn = sqlite3.connect(Config.AUDIT_DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM chat_history WHERE username = ? ORDER BY created_at ASC LIMIT ?",
            (username, limit)
        ).fetchall()
        conn.close()
        
        history = []
        for row in rows:
            msg = dict(row)
            if msg.get('table_data'):
                try:
                    msg['table'] = json.loads(msg['table_data'])
                except:
                    msg['table'] = None
            if msg.get('visualization'):
                try:
                    msg['visualization'] = json.loads(msg['visualization'])
                except:
                    msg['visualization'] = None
            if msg.get('metadata'):
                try:
                    msg['metadata'] = json.loads(msg['metadata'])
                except:
                    msg['metadata'] = None
            history.append(msg)
        
        return history
    except Exception as e:
        print(f"[CHAT_HISTORY] Error loading history: {e}")
        return []


def clear_chat_history(username: str):
    try:
        conn = sqlite3.connect(Config.AUDIT_DB_PATH)
        conn.execute("DELETE FROM chat_history WHERE username = ?", (username,))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[CHAT_HISTORY] Error clearing history: {e}")